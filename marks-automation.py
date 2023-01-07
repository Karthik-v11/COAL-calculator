import openpyxl
from tkinter import *
from PIL import ImageTk,Image
from tkinter import filedialog
from tkinter import ttk

# Give the location of the file
# path = "2020-21(oct-jan)_18CS744(CRY) -7th sem COAL.xlsx"
# response_path="bigdata.xlsx"

def run():
    workbook = openpyxl.load_workbook(path, data_only=True)
    # see_workbook=openpyxl.load_workbook(see_path)
    response_workbook=openpyxl.load_workbook(response_path,data_only=True)
    wb = openpyxl.Workbook()

    sheet1 = workbook.worksheets[0]
    sheet2 = wb.worksheets[0]
    response_sheet=response_workbook.worksheets[0]

    dims = {}

    # print(sheet1.max_row)
    # print(sheet1.max_column)
    # print(sheet1.cell(row=70, column=1).value)
    # print(sheet1.cell(row=sheet1.max_row, column=2).value)

    r1 = sheet1.cell(row=sheet1.max_row, column=1).value + 12
    r2 = sheet1.cell(row=sheet1.max_row, column=1).value + 12 + 2
    r4 = sheet1.cell(row=sheet1.max_row, column=1).value + 12 + 3
    r3 = sheet1.cell(row=sheet1.max_row, column=1).value + 12 + 4
    r5 = sheet1.cell(row=sheet1.max_row, column=1).value + 12 + 5
    r6 = sheet1.cell(row=sheet1.max_row, column=1).value + 12 + 6
    r7 = sheet1.cell(row=sheet1.max_row, column=1).value + 12 + 7
    r8=2
    r9 = sheet1.cell(row=sheet1.max_row, column=1).value + 12 + 8
    r10 = sheet1.cell(row=sheet1.max_row, column=1).value + 12 + 9

    for r in range(11, sheet1.max_row + 1):
        for c in range(1, 4):
            sheet2.cell(row=r, column=c).value = sheet1.cell(row=r, column=c).value

    sheet2['C1'] = "Internal Assesment (15) and Assignment (05)"
    sheet2['C8'] = "maximum marks"
    sheet2['C9'] = "Course Outcomes"
    sheet2['C10'] = "NAME                    Target->"
    sheet2['H1'] = "Semester End Exam (60)"
    sheet2['I8'] = 60
    sheet2['D9'] = "CO1"
    sheet2['E9'] = "CO2"
    sheet2['F9'] = "CO3"
    sheet2['G9'] = "CO4"
    sheet2['H9'] = "CO5"
    sheet2['I9'] = "SEE"
    # sheet2['K8'] = "CO1"
    # sheet2['K9'] = "CO2"
    # sheet2['K10'] = "CO3"
    # sheet2['K11'] = "CO4"
    # sheet2['L7'] = "PO1"
    # sheet2['M7'] = "PO2"
    # sheet2['N7'] = "PO3"
    # sheet2['O7'] = "PO4"
    # sheet2['P7'] = "PO5"
    # sheet2['Q7'] = "PO6"
    # sheet2['R7'] = "PO7"
    # sheet2['S7'] = "PO8"
    # sheet2['T7'] = "PO9"
    # sheet2['U7'] = "PO10"
    # sheet2['V7'] = "PO11"
    # sheet2['W7'] = "PO12"
    # sheet2['X7'] = "PSO1"
    # sheet2['Y7'] = "PSO2"
    # sheet2['Z7'] = "PSO3"
    # sheet2['Z7'] = "PSO3"
    # sheet2['A73'] = "Total number of students above target marks of each CO"
    sheet2.cell(row=r2,column=3).value="Total number of students above target marks of each CO"
    sheet2.cell(row=r3,column=3).value= "% of students"
    sheet2.cell(row=r4,column=3).value="Total Number of Students"
    sheet2.cell(row=r5,column=3).value="Attainment Level Internal(ALI)"
    sheet2.cell(row=r6,column=3).value="Attainment Level External(ALE)"
    sheet2.cell(row=r7,column=3).value="Indirect -CES"
    sheet2.cell(row=r9,column=3).value="Direct COAL(internal(50):external(50))"
    sheet2.cell(row=r10,column=3).value="Final Attainment (90% COAL+10% CES)"
    # sheet2['A74'] = "Total Number of Students"
    # sheet2['A75'] = "% of students"
    # sheet2['A76'] =  "Total Number of Students"
    # sheet2['A77'] = "Attainment Level External(ALE)"
    # sheet2['A78'] = "Indirect -CES"
    # sheet2['A79'] = "Direct COAL(internal(50):external(50))"
    # sheet2['A80'] = "Final Attainment (90% COAL+10% CES)"
    sheet2['I10'] = 24

    for r in range(10,response_sheet.max_row):
        for c in range(1,response_sheet.max_column):
            if response_sheet.cell(row=r,column=c).value=="Average":
                average_row=r

    c1=3
    for c in range(4,9):
        sheet2.cell(row=r7, column=c).value = 0
        if response_sheet.cell(row=average_row, column=c1).value is None:
            sheet2.cell(row=r7, column=c).value=0
        else:
            sheet2.cell(row=r7, column=c).value = round(response_sheet.cell(row=average_row, column=c1).value)
        c1+=1

    sheet2.cell(row=8, column=4).value = 0
    sheet2.cell(row=8, column=5).value = 0
    sheet2.cell(row=8, column=6).value = 0
    sheet2.cell(row=8, column=7).value = 0
    sheet2.cell(row=8, column=8).value = 0

    for c in range(1, sheet1.max_column+1):
        if sheet1.cell(row=9, column=c).value == None or type(sheet1.cell(row=9, column=c).value) == str:
            continue
        else:
            if sheet1.cell(row=10, column=c).value == "CO1":
                sheet2.cell(row=8, column=4).value += sheet1.cell(row=9, column=c).value
            elif sheet1.cell(row=10, column=c).value == "CO2":
                sheet2.cell(row=8, column=5).value += sheet1.cell(row=9, column=c).value
            elif sheet1.cell(row=10, column=c).value == "CO3":
                sheet2.cell(row=8, column=6).value += sheet1.cell(row=9, column=c).value
            elif sheet1.cell(row=10, column=c).value == "CO4":
                sheet2.cell(row=8, column=7).value += sheet1.cell(row=9, column=c).value
            elif sheet1.cell(row=10, column=c).value == "CO5":
                sheet2.cell(row=8, column=8).value += sheet1.cell(row=9, column=c).value

    for row in sheet2.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 5))
    for col, value in dims.items():
        sheet2.column_dimensions[col].width = value

    # print(type(sheet2.cell(row=6, column=9).value))
    sheet2.cell(row=12, column=4).value = 0
    sheet2.cell(row=12, column=5).value = 0
    sheet2.cell(row=12, column=6).value = 0
    sheet2.cell(row=12, column=7).value = 0
    sheet2.cell(row=12, column=8).value = 0
    # print(sheet1.cell(row=9, column=5).value)
    for c in range(4, 9):
        sheet2.cell(row=10, column=c).value = 0

    for c in range(4, 9):
        sheet2.cell(row=10, column=c).value = round(0.6 * sheet2.cell(row=8, column=c).value)

    for r in range(12, sheet1.max_row+1):
        c01 = 0
        c02 = 0
        c03 = 0
        c04 = 0
        c05 = 0
        for c in range(4, sheet1.max_column + 1):
            if sheet1.cell(row=10, column=c).value == "CO1":
                if sheet1.cell(row=r, column=c).value != None and type(sheet1.cell(row=r, column=c).value) != str:
                    c01 += sheet1.cell(row=r, column=c).value
                elif sheet1.cell(row=r, column=c+1).value != None and type(sheet1.cell(row=r, column=c+1).value) != str:
                    c01+=sheet1.cell(row=r, column=c + 1).value
            elif sheet1.cell(row=10, column=c).value == "CO2":
                if sheet1.cell(row=r, column=c).value != None and type(sheet1.cell(row=r, column=c).value) != str:
                    c02 += sheet1.cell(row=r, column=c).value
                elif sheet1.cell(row=r, column=c+1).value != None and type(sheet1.cell(row=r, column=c+1).value) != str:
                    c02+=sheet1.cell(row=r, column=c + 1).value
            elif sheet1.cell(row=10, column=c).value == "CO3":
                if sheet1.cell(row=r, column=c).value != None and type(sheet1.cell(row=r, column=c).value) != str:
                    c03 += sheet1.cell(row=r, column=c).value
                elif sheet1.cell(row=r, column=c+1).value != None and type(sheet1.cell(row=r, column=c+1).value) != str:
                    c03+=sheet1.cell(row=r, column=c + 1).value
            elif sheet1.cell(row=10, column=c).value == "CO4":
                if sheet1.cell(row=r, column=c).value != None and type(sheet1.cell(row=r, column=c).value) != str:
                    c04 += sheet1.cell(row=r, column=c).value
                elif sheet1.cell(row=r, column=c+1).value != None and type(sheet1.cell(row=r, column=c+1).value) != str:
                    c04+=sheet1.cell(row=r, column=c + 1).value
            elif sheet1.cell(row=10, column=c).value == "CO5":
                if sheet1.cell(row=r, column=c).value != None and type(sheet1.cell(row=r, column=c).value) != str:
                    c05+= sheet1.cell(row=r, column=c).value
                elif sheet1.cell(row=r, column=c+1).value != None and type(sheet1.cell(row=r, column=c+1).value) != str:
                    c05+=sheet1.cell(row=r, column=c + 1).value

        sheet2.cell(row=r, column=4).value = c01
        sheet2.cell(row=r, column=5).value = c02
        sheet2.cell(row=r, column=6).value = c03
        sheet2.cell(row=r, column=7).value = c04
        sheet2.cell(row=r, column=8).value = c05

    for r in range(12, sheet1.max_row+1):
            sheet2.cell(row=r,column=9).value=sheet1.cell(row=r,column=sheet1.max_column).value

    for c in range(4, 10):
        sheet2.cell(row=r4, column=c).value = sheet1.cell(row=sheet1.max_row, column=1).value
        # print(sheet1.cell(row=sheet1.max_row, column=1).value)

    for c in range(4, 10):
        count = 0
        for r in range(12, r1):
            if sheet2.cell(row=r, column=c).value is not None and type(sheet2.cell(row=r, column=c).value) != str and (sheet2.cell(row=10, column=c).value is not None and type(sheet2.cell(row=10, column=c).value) != str):
                if sheet2.cell(row=r, column=c).value >= sheet2.cell(row=10, column=c).value:
                    count += 1
        sheet2.cell(row=r2, column=c).value = count
        # print(count)

    for c in range(4, 10):
        sheet2.cell(row=r3, column=c).value = round((sheet2.cell(row=r2, column=c).value) / ((sheet2.cell(row=r4, column=c).value)) * 100,2)

    for c in range(4,10):
        if sheet2.cell(row=r3, column=c).value>=70:
            sheet2.cell(row=r5, column=c).value=3
        elif sheet2.cell(row=r3, column=c).value>=60 and sheet2.cell(row=r3, column=c).value<70:
            sheet2.cell(row=r5, column=c).value = 2
        elif sheet2.cell(row=r3, column=c).value>=50 and sheet2.cell(row=r3, column=c).value<60:
            sheet2.cell(row=r5, column=c).value = 1
        else:
            sheet2.cell(row=r5, column=c).value = 0

    for c in range(4,10):
        sheet2.cell(row=r6, column=c).value = sheet2.cell(row=r5, column=9).value

    for c in range(4, 9):
        sheet2.cell(row=r9, column=c).value = (((sheet2.cell(row=r5, column=c).value) + (sheet2.cell(row=r6, column=c).value)) / 2)

    for c in range(4, 9):
        sheet2.cell(row=r10, column=c).value = ((sheet2.cell(row=r7, column=c).value) * 0.1 + (sheet2.cell(row=r9, column=c).value) * 0.9)
        # print(sheet2.cell(row=r7, column=c).value)
    wb.save(path+" coal output.xlsx")
    print("Saved...")
    my_label = Label(root, text="Saved at "+path+"coal output.xlsx")
    my_label.pack()
    my_label.place(relx=0.50, rely=0.60, anchor=CENTER)



root = Tk()
root.title('Coal Calculator')
# root.iconbitmap('D:/project/ui/man.png')
root.geometry("800x640")
frame = Frame(root, width=50, height=40)
frame.pack()

image = Image.open("./COALdesign.jpg")
resize_image = image.resize((800, 140))
img = ImageTk.PhotoImage(resize_image)
label = Label(frame, image = img)
label.pack()

def open():
    global path
    root.filename = filedialog.askopenfilename(initialdir="/", title="Upload the .XLSX File",filetypes=(("xlsx files", "*.xlsx"), ("xls files", "*.xls")))
    my_label = Label(root, text="COAL file: "+root.filename)
    my_label.pack()
    my_label.place(relx=0.50, rely=0.50, anchor=CENTER)
    path=root.filename

def open_response():
    global response_path
    root.filename = filedialog.askopenfilename(initialdir="/", title="Upload the .XLSX File",filetypes=(("xlsx files", "*.xlsx"), ("xls files", "*.xls")))
    my_label = Label(root, text="Response file: "+root.filename)
    my_label.pack()
    my_label.place(relx=0.50, rely=0.53, anchor=CENTER)
    response_path=root.filename


my_btn = ttk.Button(root, text="Browse COAL File", command=open)
my_btn.pack()
my_btn.place(relx=0.50, rely=0.30, anchor=CENTER)
see_btn = ttk.Button(root, text="Browse Response File", command=open_response)
see_btn.pack()
see_btn.place(relx=0.50, rely=0.35, anchor=CENTER)
run_btn = ttk.Button(root, text="Run", command=run)
run_btn.pack()
run_btn.place(relx=0.50, rely=0.40, anchor=CENTER)
root.mainloop()